import requests
from urllib.parse import quote
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import urllib3

urllib3.disable_warnings()

# 설정
API_KEY = '206a1f4cf403b9db7af8a982b4951065571d7813'
BASE_URL = 'https://redmine.cresyn.com'
PROJECTS = {
    'hesh_anc2-bnp-skc202stka': 'HESH ANC2',
    'hesh_360-bhp-skc201stka-evo2': 'HESH EVO2',
    'hdx-2990': 'HDX 2990',
    'hdx-3004': 'HDX 3004',
    'ear-x': 'EAR-X'
}
AUTHORS = ['품질보증팀 김예지', '품질보증팀 이효빈', '품질보증팀 이충연']
STATUSES = ['Open', 'In Progress', 'Resolved', 'Closed']

# 날짜 기준
today = datetime.today().date()
date_str = today.strftime('%Y-%m-%d')

# 우선도 판단
def get_priority_label(issue):
    prio = issue['priority']['name']
    status = issue['status']['name']
    created = datetime.strptime(issue['created_on'][:10], "%Y-%m-%d").date()

    if status in ['Closed', 'Resolved']:
        return "완료", "완료된 이슈"

    score = 0
    if prio == 'A':
        score += 3
    elif prio == 'B':
        score += 2
    elif prio == 'C':
        score += 1

    days_open = (today - created).days
    if days_open > 14:
        score += 2
    elif days_open > 7:
        score += 1

    if score >= 5:
        return "🔺 매우 높음", "High 우선순위"
    elif score >= 3:
        return "⚠ 보통 이상", "검토 필요"
    else:
        return "⏳ 낮음", "Low 우선순위"

# 스타일
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
priority_fills = {
    "🔺 매우 높음": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    "⚠ 보통 이상": PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid"),
    "⏳ 낮음": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
center_align = Alignment(horizontal="center", vertical="center")

# 워크북
wb = Workbook()
wb.remove(wb.active)

for pid, pname in PROJECTS.items():
    url = f"{BASE_URL}/issues.json?project_id={quote(pid)}&limit=100&status_id=*"
    headers = {'X-Redmine-API-Key': API_KEY}

    try:
        response = requests.get(url, headers=headers, verify=False)
        response.raise_for_status()
        raw_issues = response.json().get('issues', [])
    except Exception as e:
        ws = wb.create_sheet(title=pname[:31])
        ws.append(["프로젝트 API 호출 실패", str(e)])
        continue

    issues = [i for i in raw_issues if i['author']['name'] in AUTHORS]
    ws = wb.create_sheet(title=pname[:31])
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 22

    # 제목 & 요약
    ws.append([f"SQA 이슈 우선순위 요약 보고서 ({date_str} 기준)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["[프로젝트명]", pname])
    ws.append(["총 이슈 수", f"{len(issues)}건"])
    not_resolved_count = sum(1 for i in issues if i['status']['name'] in ["Open", "In Progress"])
    ws.append(["잔여 이슈 수", f"{not_resolved_count}건"])
    ws.append([])
    ws.append(["[우선순위 이슈 추천 목록]"])
    ws.append(["이슈 ID", "제목", "등록일", "우선순위", "상태", "테스트 우선도", "추천 코멘트"])

    # 헤더 스타일
    for col in range(1, 8):
        cell = ws.cell(row=8, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 내용 작성
    priority_rows = []
    for issue in issues:
        label, comment = get_priority_label(issue)
        if label == "완료":
            continue
        row = [
            f"#{issue['id']}",
            issue['subject'],
            issue['created_on'][:10],
            issue['priority']['name'],
            issue['status']['name'],
            label,
            comment
        ]
        priority_rows.append((label, row))

    def priority_sort_key(item):
        label_order = {"🔺 매우 높음": 3, "⚠ 보통 이상": 2, "⏳ 낮음": 1}
        return label_order.get(item[0], 0)

    priority_rows.sort(reverse=True, key=priority_sort_key)

    for i, (_, row_data) in enumerate(priority_rows, start=9):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = center_align
            if j == 6:  # 테스트 우선도 색상 강조
                fill = priority_fills.get(value)
                if fill:
                    cell.fill = fill

# 저장
filename = f"SQA_Priority_Report_{today.strftime('%Y%m%d')}.xlsx"
wb.save(filename)
print(f"보고서 저장 완료: {filename}")
